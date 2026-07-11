"""人员名册的解析(从证书台账 xlsx)与读写(单行 JSONB 表 company_personnel)。

名册供投标时选派项目经理/总工/安全员等(见 personnel 选派服务)。解析按身份证号把同一
人在各 sheet(建造师/三类人员安全证/职称/八大员/特种工/养护工)的证书聚到一条记录;职称
表无身份证号,按姓名挂到已有记录。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Callable, Iterator

from psycopg2.extras import Json, RealDictCursor

from core.db import get_db_connection
from schemas.personnel import BuilderCert, PersonnelMember

_ENSURE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS company_personnel (
    id BIGSERIAL PRIMARY KEY,
    data JSONB NOT NULL DEFAULT '{}'::jsonb,
    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
)
"""

# 安全考核证按"项目经理优先"排序:B(项目负责人)是项目经理硬需求。
_SAFETY_CLASSES = ("A", "B", "C")


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_id(value: Any) -> str:
    # 身份证号:去空格,统一大写 X(末位校验码)。xlsx 里可能被截断,按原样去噪即可。
    return _text(value).replace(" ", "").upper()


def _valid_to(value: Any) -> str:
    text = _text(value)
    # "2023-12-28至2028-12-28" → 取"至"后的截止日;否则原样。
    if "至" in text:
        return text.split("至")[-1].strip()
    return text


def _iter_records(rows: list[tuple]) -> Iterator[Callable[[str], str]]:
    """逐行产出一个 ``getter(列名)`` —— 列映射在遇到含"姓名"的表头行时刷新。

    台账每个 sheet 内有多段(如建造师的一级/二级),各段列位置可能微移;遇表头就重建
    列映射,故对每段的列偏移都鲁棒。只产出"有姓名"的真数据行。
    """
    colmap: dict[str, int] = {}
    for row in rows:
        cells = [_text(c) for c in row]
        if "姓名" in cells:
            colmap = {name: i for i, name in enumerate(cells) if name}
            continue
        if not colmap:
            continue

        def getter(key: str, _cells=cells) -> str:
            index = colmap.get(key)
            if index is None or index >= len(_cells):
                return ""
            return _cells[index]

        if getter("姓名"):
            yield getter


def parse_roster_xlsx(xlsx_path: str | Path) -> list[PersonnelMember]:
    """解析人员证书台账 xlsx → 一人一条聚合记录。"""
    from openpyxl import load_workbook

    # 接受路径或文件型对象(openpyxl 两者都收;测试用 BytesIO)。
    source = xlsx_path if hasattr(xlsx_path, "read") else Path(xlsx_path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    by_id: dict[str, PersonnelMember] = {}
    by_name: dict[str, PersonnelMember] = {}

    def resolve(name: str, id_number: str = "") -> PersonnelMember | None:
        name = _text(name)
        id_number = _norm_id(id_number)
        if not name:
            return None
        if id_number and id_number in by_id:
            member = by_id[id_number]
            if not member.id_number:
                member.id_number = id_number
            return member
        if not id_number and name in by_name:
            return by_name[name]
        member = PersonnelMember(name=name, id_number=id_number)
        if id_number:
            by_id[id_number] = member
        by_name.setdefault(name, member)
        return member

    def rows_of(sheet: str) -> list[tuple]:
        if sheet not in workbook.sheetnames:
            return []
        return list(workbook[sheet].iter_rows(values_only=True))

    # 建造师(=项目经理候选资格):类别列直接给"一级建造师/二级建造师",专业列给专业。
    for getter in _iter_records(rows_of("正奇建造师证书")):
        category = getter("类别")
        if "建造师" not in category:
            continue  # 跳过同表里的造价工程师等
        member = resolve(getter("姓名"), getter("身份证号"))
        if member is None:
            continue
        member.builder_certs.append(
            BuilderCert(
                level=category,
                specialty=getter("专业"),
                cert_no=getter("证书编号"),
                valid_to=_valid_to(getter("有效期")),
            )
        )

    # 三类人员安全考核证:类别 A/B/C。项目经理须 B 证。
    # ⚠️台账此表按 A/B/C 分块,"类别"是合并单元格——只有每块第一行有值,其余行读出来
    # 是空。必须"沿用上一行的类别"(forward-fill),否则 64 行只认得 3 行(每块首行),
    # 名册几乎人人"缺安全B证"(2026-07-12 用户实测揪出)。
    current_class = ""
    for getter in _iter_records(rows_of("三类人员证书")):
        raw = getter("类别").upper()
        if raw in _SAFETY_CLASSES:
            current_class = raw  # 块首行:更新当前类别
        elif raw:
            # 非A/B/C的杂值(如段标题"交安"串进类别列):不更新也不沿用,防污染
            continue
        if not current_class:
            continue
        member = resolve(getter("姓名"), getter("身份证号"))
        if member is None:
            continue
        if current_class not in member.safety_cert_classes:
            member.safety_cert_classes.append(current_class)
        member.safety_cert_no = member.safety_cert_no or getter("证书编号")

    # 八大员(新/旧):证书类别 = 标准员/材料员/…
    for sheet in ("八大员 (新)", "八大员（旧）"):
        for getter in _iter_records(rows_of(sheet)):
            member = resolve(getter("姓名"), getter("身份证号"))
            role = getter("证书类别")
            if member and role and role not in member.eight_roles:
                member.eight_roles.append(role)

    # 特种工:工种
    for getter in _iter_records(rows_of("特种工 (新)")):
        member = resolve(getter("姓名"), getter("身份证号"))
        work = getter("工种")
        if member and work and work not in member.special_works:
            member.special_works.append(work)

    # 养护工:技能等级
    for getter in _iter_records(rows_of("养护工")):
        member = resolve(getter("姓名"), getter("身份证号"))
        grade = getter("技能等级")
        if member and grade:
            member.special_works.append(f"公路养护工{grade}")

    # 职称(工程师证书):该 sheet 无身份证号且是"左右两栏"布局,每行含两人,按姓名挂。
    for row in rows_of("正奇工程师证书"):
        cells = [_text(c) for c in row]
        for base in (0, 9):  # 左栏 0-6、右栏 9-15
            if base >= len(cells):
                continue
            name = cells[base]
            if not name or name == "姓名":
                continue
            category = cells[base + 2] if base + 2 < len(cells) else ""
            if "工程师" not in category:
                continue
            member = resolve(name)
            if member and not member.title:
                member.title = category
                member.title_specialty = cells[base + 1] if base + 1 < len(cells) else ""

    workbook.close()
    # 按"是否项目经理候选 + 姓名"稳定排序,候选靠前
    members = list({id(m): m for m in [*by_id.values(), *by_name.values()]}.values())
    members.sort(key=lambda m: (not m.is_pm_candidate, m.name))
    return members


_BUILDER_LEVELS = {"一级建造师证": "一级建造师", "二级建造师证": "二级建造师"}
_KNOWN_SPECIALTIES = (
    "公路", "市政", "建筑", "机电", "水利", "铁路", "港口", "通信", "矿业", "机场",
)


def _clean_name(name: str) -> str:
    # 知识库 owner_name 偶有文件名残渣(尾部"照/_2"等);轻清洗,接受少量噪声。
    name = _text(name)
    name = re.sub(r"[_\-\s]?\d+$", "", name)  # 去尾部 _2 / -1
    return name.strip()


def kb_builder_candidates() -> dict[str, list[BuilderCert]]:
    """从知识库人员证件抽建造师候选(姓名 → 建造师证列表)。

    知识库是历年全量(含离职/过期/重复),证书等级在 certificate_type,专业编码在文件名
    结构 ``…_X级建造师证_<专业>_…``。按姓名聚合,去重 (等级,专业)。
    """
    spec_re = re.compile(r"建造师证_([^_]+)")
    candidates: dict[str, set[tuple[str, str]]] = {}
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT DISTINCT d.metadata_json->>'owner_name' AS name,
                       d.metadata_json->>'certificate_type' AS cert_type,
                       d.file_name
                FROM documents d
                WHERE d.project_id IS NULL
                  AND d.metadata_json->>'owner_name' <> ''
                  AND (d.file_name ILIKE '%建造师%'
                       OR d.metadata_json->>'certificate_type' ILIKE '%建造师%')
                """
            )
            rows = cursor.fetchall()
    for name, cert_type, file_name in rows:
        name = _clean_name(name)
        level = _BUILDER_LEVELS.get(_text(cert_type), "")
        if not name or not level:
            continue
        match = spec_re.search(_text(file_name))
        specialty = match.group(1) if match else ""
        if specialty == "通用" or not any(k in specialty for k in _KNOWN_SPECIALTIES):
            specialty = ""  # 通用/无法识别 → 留空(匹配时按"未注明专业"处理)
        candidates.setdefault(name, set()).add((level, specialty))
    return {
        name: [BuilderCert(level=lv, specialty=sp) for lv, sp in sorted(pairs)]
        for name, pairs in candidates.items()
    }


def merge_kb_builders(
    members: list[PersonnelMember], kb_candidates: dict[str, list[BuilderCert]]
) -> list[PersonnelMember]:
    """把知识库建造师候选并入名册:台账已有的人优先用台账(更干净),台账没有的人新增
    (标 source=知识库,提示用户其可能离职/过期/需核验)。返回合并后的名册。"""
    by_name = {m.name: m for m in members}
    for name, certs in kb_candidates.items():
        existing = by_name.get(name)
        if existing is None:
            members.append(
                PersonnelMember(name=name, builder_certs=certs, source="知识库")
            )
        elif not existing.builder_certs:
            # 台账里这人无建造师证 → 用知识库补,标双来源。
            existing.builder_certs = certs
            existing.source = "台账+知识库"
    members.sort(key=lambda m: (not m.is_pm_candidate, m.name))
    return members


def _absorb(host: PersonnelMember, other: PersonnelMember) -> None:
    """把脏名变体 other 的证书并进真人 host(去重)。"""
    seen = {(c.level, c.specialty) for c in host.builder_certs}
    for cert in other.builder_certs:
        if (cert.level, cert.specialty) not in seen:
            host.builder_certs.append(cert)
            seen.add((cert.level, cert.specialty))
    for cls in other.safety_cert_classes:
        if cls not in host.safety_cert_classes:
            host.safety_cert_classes.append(cls)
    host.safety_cert_no = host.safety_cert_no or other.safety_cert_no
    if not host.title and other.title:
        host.title, host.title_specialty = other.title, other.title_specialty
    host.eight_roles = sorted(set(host.eight_roles) | set(other.eight_roles))
    host.special_works = sorted(set(host.special_works) | set(other.special_works))
    host.id_number = host.id_number or other.id_number
    sources = host.source + other.source
    if "台账" in sources and "知识库" in sources:
        host.source = "台账+知识库"


def dedupe_roster(members: list[PersonnelMember]) -> list[PersonnelMember]:
    """合并 OCR 脏名拆出来的同一人。

    脏名特征=真名(2~3字)+1个垃圾尾字(执/返/照/市/退/签…)。两条保守规则,只动 ≥4 字的
    长名(中文真名极少 4 字),不碰 2↔3 字以免误并真人:
    1. 长名(≥4字)的某个 ≥2字前缀正好是已存在的更短真名 → 脏变体,并入(康白华执→康白华)。
    2. ≥2 个 ≥4字 名共享同一(去末字)前缀 → 都是该前缀真名的脏变体,归并(夏冬梅照/签→夏冬梅)。
    """
    ordered = sorted(members, key=lambda m: (len(m.name), m.name))
    kept: list[PersonnelMember] = []
    for member in ordered:
        host = None
        if len(member.name) >= 4:
            for candidate in kept:
                if (
                    2 <= len(candidate.name) < len(member.name)
                    and member.name.startswith(candidate.name)
                ):
                    host = candidate
                    break
        if host is not None:
            _absorb(host, member)
        else:
            kept.append(member)

    # 规则2:同(去末字)前缀的多个长脏名 → 归到该前缀真名
    from collections import defaultdict

    buckets: dict[str, list[PersonnelMember]] = defaultdict(list)
    for member in list(kept):
        if len(member.name) >= 4:
            buckets[member.name[:-1]].append(member)
    for prefix, group in buckets.items():
        if len(group) >= 2 and len(prefix) >= 2:
            host = group[0]
            host.name = prefix
            for member in group[1:]:
                _absorb(host, member)
                kept.remove(member)

    kept.sort(key=lambda m: (not m.is_pm_candidate, m.name))
    return kept


def get_personnel_roster() -> dict[str, Any]:
    """返回保存的名册(list[PersonnelMember dict]) + 更新时间;空库不报错。"""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(_ENSURE_TABLE_SQL)
            conn.commit()
            cursor.execute(
                "SELECT data, updated_at FROM company_personnel ORDER BY id LIMIT 1"
            )
            row = cursor.fetchone()
    data = dict(row["data"]) if row and row.get("data") else {}
    return {
        "roster": data.get("roster", []),
        "updated_at": row["updated_at"].isoformat() if row else None,
    }


def save_personnel_roster(members: list[PersonnelMember]) -> dict[str, Any]:
    payload = {"roster": [m.model_dump() for m in members]}
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(_ENSURE_TABLE_SQL)
            cursor.execute("SELECT id FROM company_personnel ORDER BY id LIMIT 1")
            row = cursor.fetchone()
            if row:
                cursor.execute(
                    "UPDATE company_personnel SET data = %s, updated_at = NOW() "
                    "WHERE id = %s",
                    (Json(payload), row["id"]),
                )
            else:
                cursor.execute(
                    "INSERT INTO company_personnel (data) VALUES (%s)",
                    (Json(payload),),
                )
            conn.commit()
    return get_personnel_roster()
