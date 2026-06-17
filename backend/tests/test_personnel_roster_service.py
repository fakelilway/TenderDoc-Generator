from io import BytesIO

from openpyxl import Workbook

from schemas.personnel import BuilderCert, PersonnelMember
from services.personnel_roster_service import merge_kb_builders, parse_roster_xlsx


def _build_xlsx() -> BytesIO:
    """模拟人员证书台账:建造师(多段)、三类人员安全证、工程师职称(左右两栏、无身份证号)。"""
    wb = Workbook()
    wb.remove(wb.active)

    jz = wb.create_sheet("正奇建造师证书")
    jz.append(["一级建造师"])
    jz.append(["序号", "姓名", "身份证号", "", "证书编号", "学历", "类别", "专业", "发证日期", "有效期"])
    jz.append(["1", "江舟", "3401041970", "", "皖1342006", "专科", "一级建造师", "公路工程", "2023", "2023至2028"])
    jz.append(["二级建造师"])
    jz.append(["序号", "姓名", "身份证号", "资格证书编号", "证书编号", "学历", "类别", "专业", "发证日期", "有效期"])
    jz.append(["1", "李刚", "3425231988", "x", "皖234", "专科", "二级建造师", "市政公用工程", "2025", "2025至2030"])
    # 同表里的造价工程师不该算建造师候选
    jz.append(["注册造价工程师"])
    jz.append(["序号", "姓名", "身份证号", "x", "证书编号", "学历", "类别", "专业", "发证日期", "有效期"])
    jz.append(["1", "王五", "3401231990", "y", "z", "本科", "二级造价工程师", "土建", "2024", "2029"])

    sc = wb.create_sheet("三类人员证书")
    sc.append(["交安"])
    sc.append([])
    sc.append(["类别", "序号", "姓名", "身份证号", "证书编号", "", "发证日期", "有效期"])
    sc.append(["B", "1", "李刚", "3425231988", "皖交安B17", "", "2017", "2027"])  # 李刚 有B证

    gc = wb.create_sheet("正奇工程师证书")
    # 左右两栏、无身份证号,按姓名挂职称
    gc.append(["姓名", "专业", "类别", "发证日期", "证书编号", "社保", "备注", "", "", "姓名", "专业", "类别", "发证日期", "证书编号", "社保", "备注"])
    gc.append(["江舟", "交通", "高级工程师", "2006", "060198", "", "", "", "", "李刚", "市政", "工程师", "2010", "0700", "", ""])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_parse_roster_joins_certs_by_id_and_name() -> None:
    members = parse_roster_xlsx(_build_xlsx())
    by_name = {m.name: m for m in members}

    # 江舟:一级公路建造师 + 高级工程师职称(职称表无身份证号,按姓名挂上)
    jiang = by_name["江舟"]
    assert jiang.is_pm_candidate
    assert jiang.builder_levels == ["一级建造师"]
    assert jiang.builder_specialties == ["公路工程"]
    assert jiang.title == "高级工程师"

    # 李刚:二级市政建造师 + 安全B证(按身份证号串起来)+ 工程师职称
    li = by_name["李刚"]
    assert li.is_pm_candidate
    assert li.builder_specialties == ["市政公用工程"]
    assert "B" in li.safety_cert_classes
    assert li.title == "工程师"

    # 只有建造师才是项目经理候选:造价工程师(王五)不进候选(本例他无其它证→不入名册)
    pm_names = [m.name for m in members if m.is_pm_candidate]
    assert set(pm_names) == {"江舟", "李刚"}
    assert "王五" not in by_name  # 造价-only,未被任一证书表收录


def test_merge_kb_builders_supplements_and_adds() -> None:
    """知识库建造师并入:台账已有的保台账;台账有人但无证→KB补(双来源);台账没的→新增。"""
    roster = [
        PersonnelMember(
            name="江舟",
            builder_certs=[BuilderCert(level="一级建造师", specialty="公路工程")],
            source="台账",
        ),
        PersonnelMember(name="许明英", source="台账"),  # 台账有人、无建造师证
    ]
    kb = {
        "江舟": [BuilderCert(level="一级建造师", specialty="公路工程")],
        "许明英": [BuilderCert(level="二级建造师", specialty="市政公用工程")],
        "周明明": [BuilderCert(level="二级建造师", specialty="公路工程")],
    }
    merged = {m.name: m for m in merge_kb_builders(roster, kb)}

    assert merged["江舟"].source == "台账"  # 已是候选,保台账不动
    assert merged["许明英"].is_pm_candidate
    assert merged["许明英"].source == "台账+知识库"  # 用 KB 补证
    assert merged["周明明"].source == "知识库"  # 台账没这人 → 新增
    assert merged["周明明"].builder_specialties == ["公路工程"]


def test_dedupe_roster_merges_dirty_name_variants() -> None:
    """OCR 脏名拆出的同一人合并:康白华执/返→康白华、夏冬梅照/签→夏冬梅;真人不误并。"""
    from services.personnel_roster_service import dedupe_roster

    def cert():
        return [BuilderCert(level="二级建造师", specialty="公路工程")]

    members = [
        PersonnelMember(name="康白华", builder_certs=cert(), source="台账+知识库"),
        PersonnelMember(name="康白华执", builder_certs=cert(), source="知识库"),
        PersonnelMember(name="康白华返", builder_certs=cert(), source="知识库"),
        PersonnelMember(name="夏冬梅照", builder_certs=cert(), source="知识库"),
        PersonnelMember(name="夏冬梅签", builder_certs=cert(), source="知识库"),
        PersonnelMember(
            name="江舟",
            builder_certs=[BuilderCert(level="一级建造师", specialty="公路工程")],
            source="台账",
        ),
    ]
    out = {m.name for m in dedupe_roster(members)}
    assert out == {"康白华", "夏冬梅", "江舟"}

    # 真人不误并:王俊(2字) vs 王俊明(3字)各自保留(规则只动 ≥4 字脏名)
    pair = dedupe_roster(
        [
            PersonnelMember(name="王俊", builder_certs=cert()),
            PersonnelMember(name="王俊明", builder_certs=cert()),
        ]
    )
    assert {m.name for m in pair} == {"王俊", "王俊明"}
