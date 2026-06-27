"""工程量清单 Excel(.xlsx)抽取测试:把表格行序列化成文本,供 build_boq 估占比。"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from utils.file_parser import (
    SUPPORTED_EXTENSIONS,
    extract_text,
    extract_text_from_xlsx,
)


def _boq_xlsx_bytes() -> bytes:
    """仿造价软件导出的工程量清单:表头 + 两条分部分项 + 一条全空行。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "分部分项工程量清单"
    ws.append(["序号", "项目编码", "项目名称", "计量单位", "工程数量", "综合单价"])
    ws.append([1, "040101001", "路基填方", "m³", 126000, 35.5])
    ws.append([2, "040203001", "沥青混凝土路面", "m²", 82000, 120.0])
    ws.append([None, None, None, None, None, None])  # 全空行
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_in_supported_extensions() -> None:
    assert {".xlsx", ".xlsm", ".xls"} <= SUPPORTED_EXTENSIONS


def test_extract_text_from_xlsx_serialises_rows() -> None:
    text = extract_text_from_xlsx(_boq_xlsx_bytes())
    # 每条工程量项与其数量都进了文本,列之间以 " | " 分隔。
    assert "路基填方" in text and "126000" in text
    assert "沥青混凝土路面" in text and "82000" in text
    assert " | " in text
    # 全空行被跳过:只有表头 + 两条数据 = 3 行。
    assert len(text.splitlines()) == 3


def test_extract_text_routes_xlsx_by_suffix() -> None:
    text = extract_text(_boq_xlsx_bytes(), filename="工程量清单.xlsx")
    assert "路基填方" in text


def test_extract_text_routes_xlsx_by_content_type() -> None:
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    text = extract_text(_boq_xlsx_bytes(), filename="boq", content_type=ct)
    assert "沥青混凝土路面" in text


def test_multi_sheet_gets_sheet_headers() -> None:
    wb = Workbook()
    wb.active.title = "路基"
    wb.active.append(["路基填方", 126000])
    second = wb.create_sheet("路面")
    second.append(["沥青混凝土", 82000])
    buf = BytesIO()
    wb.save(buf)
    text = extract_text_from_xlsx(buf.getvalue())
    assert "## 工作表:路基" in text and "## 工作表:路面" in text
