"""Import the company performance ledger (历年中标项目统计表) into the knowledge base.

业绩(track record) is a scored/qualification item in most construction tenders
(类似业绩). The useful form is NOT the 2000+ scanned 中标/合同/验收 images — it is the
structured ledger: one record per won project with 名称/金额/工期/类型/日期, so the
system can match a tender's 类似业绩 requirement and fill 业绩汇总表 blanks.

This parses the consolidated '汇总表' sheet (one consistent schema, ~253 projects)
and ingests one structured, RAG-searchable record per project. 中标金额 has no
dedicated metadata field in the KB, so it is written into the searchable text AND
encoded as an amount-band tag for coarse "合同额≥X" filtering; project_year /
project_type are first-class filterable metadata.

用法:
  预览(不写库):  PYTHONPATH=backend .venv/bin/python backend/scripts/import_performance_ledger.py <xlsx> --out preview.csv
  入库:          ... --import-to-kb --import-report report.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

COMPANY_NAME = "安徽正奇建设有限公司"
LEDGER_TAG = "导入批次-业绩台账"
DEFAULT_SHEET = "汇总表"
DATA_START_ROW = 4  # rows 1-3 are the merged title/header block

# Column indices in the 汇总表 sheet (0-based).
COL = {
    "contract_no": 1,
    "project_name": 2,
    "project_manager": 3,
    "chief_engineer": 4,
    "bid_date": 5,
    "sign_date": 6,
    "start_date": 7,
    "complete_date": 8,
    "handover_date": 9,
    "finish_date": 10,
    "duration": 11,
    "amount": 12,
    "summary": 13,
    "scan_award": 14,
    "scan_contract": 15,
    "scan_handover": 16,
    "owner": 19,
}

_EMPTY = {None, "", "/", "·", "无", "—", "-"}

# (keyword..., type) — ordered by specificity so a "农村公路安全生命防护" project is
# tagged 交安工程 (its scoring lens) before the generic 公路工程.
_TYPE_RULES = (
    (("隧道",), "隧道工程"),
    (("桥", "危桥"), "桥梁工程"),
    (("安全生命防护", "安防", "护栏", "标志标线", "交通安全设施", "波形梁"), "交安工程"),
    (("养护", "大修", "中修", "预防性养护", "维修"), "公路养护"),
    (("管网", "排水", "污水", "给水", "市政", "照明", "管廊"), "市政工程"),
    (("小区", "整治", "粉刷", "房屋", "房建", "学校", "运动场", "楼", "安置", "厂房"), "房屋建筑"),
    (("公路", "道路", "路面", "路基", "国道", "省道", "县道", "乡道", "村道", "通村", "通组", "组组通", "公路工程"), "公路工程"),
)


@dataclass
class PerformanceRecord:
    contract_no: str = ""
    project_name: str = ""
    project_manager: str = ""
    chief_engineer: str = ""
    bid_date: str = ""
    sign_date: str = ""
    duration: str = ""
    amount_raw: str = ""
    amount_wan: float | None = None
    amount_band: str = ""
    summary: str = ""
    owner: str = ""
    project_type: str = ""
    project_year: int | None = None
    has_award_scan: bool = False
    has_contract_scan: bool = False
    has_handover_scan: bool = False
    suggested_filename: str = ""


def _cell(row: tuple, key: str) -> object:
    index = COL[key]
    return row[index] if index < len(row) else None


def _text(value: object) -> str:
    if value in _EMPTY:
        return ""
    return str(value).replace("\n", " ").strip()


def _has_scan(value: object) -> bool:
    return _text(value) not in {"", "无"}


def _parse_amount(value: object) -> float | None:
    """Return the contract amount in 万元 (1万 = 10,000), or None.

    Cells are usually a number (元) but can be a wrapped string like
    '1798910.9\\n1041'; take the first numeric token and convert to 万.
    """
    if value in _EMPTY:
        return None
    if isinstance(value, (int, float)):
        yuan = float(value)
    else:
        first = str(value).splitlines()[0].replace(",", "").replace("，", "")
        match = re.search(r"\d+(?:\.\d+)?", first)
        if not match:
            return None
        yuan = float(match.group())
        if "万" in str(value):
            return round(yuan, 2)
    return round(yuan / 10000.0, 2)


def _amount_band(wan: float | None) -> str:
    if wan is None:
        return "金额-未知"
    if wan < 500:
        return "金额-500万以下"
    if wan < 1000:
        return "金额-500至1000万"
    if wan < 3000:
        return "金额-1000至3000万"
    if wan < 10000:
        return "金额-3000万至1亿"
    return "金额-1亿以上"


def _extract_year(*values: object) -> int | None:
    for value in values:
        match = re.search(r"(20\d{2})", str(value or ""))
        if match:
            return int(match.group(1))
    return None


def _infer_type(name: str, summary: str) -> str:
    """Return the single most-specific project type (rules ordered specific→generic).

    A single primary type avoids false-positive 类似业绩 matches — e.g. a 公路养护
    project must not also count as a new-build 公路工程.
    """
    text = f"{name} {summary}"
    for keywords, project_type in _TYPE_RULES:
        if any(keyword in text for keyword in keywords):
            return project_type
    return "其他工程"


def _safe_segment(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\x00-\x1f]+", "", value).strip()
    return re.sub(r"\s+", "", cleaned)[:30]


def parse_ledger(xlsx_path: str | Path, sheet_name: str = DEFAULT_SHEET) -> list[PerformanceRecord]:
    # 惰性导入:金额/类型规则被 import_similar_project_info 复用,那条链不碰 xlsx,
    # 不能因为顶层 import openpyxl 把只装了 python-docx 的环境拖崩。
    from openpyxl import load_workbook

    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在;可选: {workbook.sheetnames}")
    worksheet = workbook[sheet_name]

    records: list[PerformanceRecord] = []
    used_names: set[str] = set()
    for row in worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        name = _text(_cell(row, "project_name"))
        if not name:
            continue
        summary = _text(_cell(row, "summary"))
        project_type = _infer_type(name, summary)
        amount_wan = _parse_amount(_cell(row, "amount"))
        year = _extract_year(
            _cell(row, "bid_date"), _cell(row, "sign_date"), _cell(row, "start_date")
        )
        record = PerformanceRecord(
            contract_no=_text(_cell(row, "contract_no")),
            project_name=name,
            project_manager=_text(_cell(row, "project_manager")),
            chief_engineer=_text(_cell(row, "chief_engineer")),
            bid_date=_text(_cell(row, "bid_date")),
            sign_date=_text(_cell(row, "sign_date")),
            duration=_text(_cell(row, "duration")),
            amount_raw=_text(_cell(row, "amount")),
            amount_wan=amount_wan,
            amount_band=_amount_band(amount_wan),
            summary=summary,
            owner=_text(_cell(row, "owner")),
            project_type=project_type,
            project_year=year,
            has_award_scan=_has_scan(_cell(row, "scan_award")),
            has_contract_scan=_has_scan(_cell(row, "scan_contract")),
            has_handover_scan=_has_scan(_cell(row, "scan_handover")),
        )
        base = f"业绩_{year or '年份待核'}_{_safe_segment(name)}"
        candidate = f"{base}.txt"
        index = 2
        while candidate in used_names:
            candidate = f"{base}_{index}.txt"
            index += 1
        used_names.add(candidate)
        record.suggested_filename = candidate
        records.append(record)
    return records


def build_record_text(record: PerformanceRecord) -> str:
    amount = f"{record.amount_wan}万元" if record.amount_wan is not None else "待核"
    scans = (
        f"中标{'有' if record.has_award_scan else '无'} "
        f"合同{'有' if record.has_contract_scan else '无'} "
        f"交工{'有' if record.has_handover_scan else '无'}"
    )
    lines = [
        f"业绩项目：{record.project_name}",
        f"项目类型：{record.project_type}",
        f"中标金额：{amount}",
        f"中标日期：{record.bid_date or '待核'}" + (f"（{record.project_year}年）" if record.project_year else ""),
        f"工期：{record.duration or '待核'}",
        f"项目经理：{record.project_manager or '待核'}　项目总工：{record.chief_engineer or '待核'}",
        f"发包人：{record.owner}" if record.owner else "",
        f"工程内容：{record.summary}" if record.summary else "",
        f"合同编号：{record.contract_no}" if record.contract_no else "",
        f"原件可得：{scans}",
    ]
    return "\n".join(line for line in lines if line)


def _record_tags(record: PerformanceRecord) -> list[str]:
    tags = ["业绩", record.amount_band, LEDGER_TAG, record.project_type]
    if record.project_year:
        tags.append(f"{record.project_year}年")
    seen: set[str] = set()
    return [t for t in tags if t and not (t in seen or seen.add(t))]


def import_records_to_kb(records: list[PerformanceRecord]) -> list[dict[str, object]]:
    from services import knowledge_service

    results: list[dict[str, object]] = []
    for record in records:
        text = build_record_text(record)
        try:
            indexed = knowledge_service.index_uploaded_knowledge(
                file_bytes=text.encode("utf-8"),
                filename=record.suggested_filename,
                content_type="text/plain",
                project_type=record.project_type,
                document_type="业绩记录",
                document_category="业绩",
                volume="资格文件",
                project_year=record.project_year,
                owner_type="公司",
                owner_name=COMPANY_NAME,
                usage_scope="可用于投标",
                verified_status="待核验",
                image_insertable=False,
                tags=_record_tags(record),
                ingestion_mode="rag_text",
            )
            results.append(
                {
                    "project_name": record.project_name,
                    "suggested_filename": record.suggested_filename,
                    "status": "imported",
                    "document_id": indexed.get("document_id"),
                    "indexing_status": indexed.get("indexing_status"),
                    "message": "",
                }
            )
        except Exception as error:  # noqa: BLE001 - per-row isolation, like the cert importer
            results.append(
                {
                    "project_name": record.project_name,
                    "suggested_filename": record.suggested_filename,
                    "status": "failed",
                    "document_id": "",
                    "indexing_status": "",
                    "message": str(error),
                }
            )
    return results


def write_preview_csv(records: list[PerformanceRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(asdict(PerformanceRecord()).keys())
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def write_import_report(results: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["project_name", "suggested_filename", "status", "document_id", "indexing_status", "message"]
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow({key: result.get(key, "") for key in fieldnames})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse and optionally import the performance ledger.")
    parser.add_argument("xlsx", type=Path, help="历年中标项目统计表 .xlsx path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet to parse (default 汇总表)")
    parser.add_argument("--out", type=Path, required=True, help="Preview CSV output path")
    parser.add_argument("--import-to-kb", action="store_true", help="Import records into the knowledge base")
    parser.add_argument("--import-report", type=Path, help="CSV report for import results")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    records = parse_ledger(args.xlsx.expanduser().resolve(), sheet_name=args.sheet)
    write_preview_csv(records, args.out.expanduser().resolve())
    print(f"Parsed {len(records)} performance records")
    print(f"Preview CSV: {args.out.expanduser().resolve()}")
    if args.import_to_kb:
        results = import_records_to_kb(records)
        report_path = (
            args.import_report.expanduser().resolve()
            if args.import_report
            else args.out.expanduser().resolve().with_name("performance_import_report.csv")
        )
        write_import_report(results, report_path)
        imported = sum(1 for r in results if r["status"] == "imported")
        failed = sum(1 for r in results if r["status"] == "failed")
        print(f"Import: {imported} imported, {failed} failed")
        print(f"Import report: {report_path}")


if __name__ == "__main__":
    main()
